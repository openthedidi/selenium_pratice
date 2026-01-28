import openpyxl

file = "D:\\selenium-driver\\ClassExamples\\ClassExamples\\testdata.xlsx"


workbook = openpyxl.load_workbook(file)
sheet = workbook["測試資料分頁1"]


# if only sheet can use .active
# heet = workbook.active

rows = sheet.max_row
cols = sheet.max_column

print(f"rows: {rows}, cols: {cols}")

# ---以row為單位
for row in range(2, 3):
    for col in range(1, cols+1):
        sheet.cell(row, col).value = (f'test:' + str(rows))

workbook.save(file)
workbook.close()
