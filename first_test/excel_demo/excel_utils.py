import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side


def getRowCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return (sheet.max_row)


def getColumnCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return (sheet.max_column)


def readData(file, sheetName, rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(rownum, columnno).value


def writeData(file, sheetName, rownum, columnno, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(rownum, columnno).value = data
    workbook.save(file)


def fillGreenColor(file, sheetName, rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    greenFill = PatternFill(start_color='60b212',
                            end_color='60b212',
                            fill_type='solid')
    sheet.cell(rownum, columnno).fill = greenFill
    workbook.save(file)


def fillRedColor(file, sheetName, rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    redFill = PatternFill(start_color='ff0000',
                          end_color='ff0000',
                          fill_type='solid')
    sheet.cell(rownum, columnno).fill = redFill
    workbook.save(file)


# ========== 以下為新增的 10 個常用方法 ==========

def createWorkbook(file, sheetName="Sheet1"):
    """建立新的 Excel 檔案"""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheetName
    workbook.save(file)


def getSheetNames(file):
    """取得所有工作表名稱"""
    workbook = openpyxl.load_workbook(file)
    return workbook.sheetnames


def createSheet(file, sheetName):
    """新增工作表"""
    workbook = openpyxl.load_workbook(file)
    workbook.create_sheet(sheetName)
    workbook.save(file)


def deleteSheet(file, sheetName):
    """刪除工作表"""
    workbook = openpyxl.load_workbook(file)
    del workbook[sheetName]
    workbook.save(file)


def readRow(file, sheetName, rownum):
    """讀取整行資料，回傳 list"""
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    row_data = []
    for col in range(1, sheet.max_column + 1):
        row_data.append(sheet.cell(rownum, col).value)
    return row_data


def readColumn(file, sheetName, columnno):
    """讀取整欄資料，回傳 list"""
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    col_data = []
    for row in range(1, sheet.max_row + 1):
        col_data.append(sheet.cell(row, columnno).value)
    return col_data


def writeRow(file, sheetName, rownum, data_list):
    """寫入整行資料，data_list 為 list"""
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    for col, value in enumerate(data_list, start=1):
        sheet.cell(rownum, col).value = value
    workbook.save(file)


def insertRow(file, sheetName, rownum):
    """在指定位置插入一行"""
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.insert_rows(rownum)
    workbook.save(file)


def deleteRow(file, sheetName, rownum):
    """刪除指定行"""
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.delete_rows(rownum)
    workbook.save(file)


def mergeCells(file, sheetName, start_row, start_col, end_row, end_col):
    """合併儲存格"""
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.merge_cells(start_row=start_row, start_column=start_col,
                      end_row=end_row, end_column=end_col)
    workbook.save(file)
