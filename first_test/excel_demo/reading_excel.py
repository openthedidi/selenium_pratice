import openpyxl


# 讀取順序會是:file > workbook > sheets > rows > cell
file = "D:\\selenium-driver\\ClassExamples\\ClassExamples\\data.xlsx"

workbook = openpyxl.load_workbook(file)
sheet = workbook["Sheet1"]

rows = sheet.max_row
cols = sheet.max_column


print(f"rows: {rows}, cols: {cols}")

# -----------read all data-------------
for row in range(1, rows+1):
    for col in range(1, cols+1):
        # print(f"row: {row}, col: {col}")
        # print(sheet.cell(row=row, column=col).value)

        # 仿表格的方式print
        print(sheet.cell(row, col).value, end="    ")
    print()
