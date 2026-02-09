import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side


def get_row_count(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.max_row


def get_column_count(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.max_column


def read_data(file, sheet_name, row_num, column_no):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.cell(row_num, column_no).value


def write_data(file, sheet_name, row_num, column_no, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    sheet.cell(row_num, column_no).value = data
    workbook.save(file)


def fill_green_color(file, sheet_name, row_num, column_no):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    green_fill = PatternFill(start_color='60b212',
                             end_color='60b212',
                             fill_type='solid')
    sheet.cell(row_num, column_no).fill = green_fill
    workbook.save(file)


def fill_red_color(file, sheet_name, row_num, column_no):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    red_fill = PatternFill(start_color='ff0000',
                           end_color='ff0000',
                           fill_type='solid')
    sheet.cell(row_num, column_no).fill = red_fill
    workbook.save(file)


# ========== 以下為新增的 10 個常用方法 ==========

def create_workbook(file, sheet_name="Sheet1"):
    """建立新的 Excel 檔案"""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    workbook.save(file)


def get_sheet_names(file):
    """取得所有工作表名稱"""
    workbook = openpyxl.load_workbook(file)
    return workbook.sheetnames


def create_sheet(file, sheet_name):
    """新增工作表"""
    workbook = openpyxl.load_workbook(file)
    workbook.create_sheet(sheet_name)
    workbook.save(file)


def delete_sheet(file, sheet_name):
    """刪除工作表"""
    workbook = openpyxl.load_workbook(file)
    del workbook[sheet_name]
    workbook.save(file)


def read_row(file, sheet_name, row_num):
    """讀取整行資料，回傳 list"""
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    row_data = []
    for col in range(1, sheet.max_column + 1):
        row_data.append(sheet.cell(row_num, col).value)
    return row_data


def read_column(file, sheet_name, column_no):
    """讀取整欄資料，回傳 list"""
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    col_data = []
    for row in range(1, sheet.max_row + 1):
        col_data.append(sheet.cell(row, column_no).value)
    return col_data


def write_row(file, sheet_name, row_num, data_list):
    """寫入整行資料，data_list 為 list"""
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    for col, value in enumerate(data_list, start=1):
        sheet.cell(row_num, col).value = value
    workbook.save(file)


def insert_row(file, sheet_name, row_num):
    """在指定位置插入一行"""
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    sheet.insert_rows(row_num)
    workbook.save(file)


def delete_row(file, sheet_name, row_num):
    """刪除指定行"""
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    sheet.delete_rows(row_num)
    workbook.save(file)


def merge_cells(file, sheet_name, start_row, start_col, end_row, end_col):
    """合併儲存格"""
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    sheet.merge_cells(start_row=start_row, start_column=start_col,
                      end_row=end_row, end_column=end_col)
    workbook.save(file)
